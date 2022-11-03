# -*- coding: utf-8 -*-
import xlsxwriter
import xlrd
from openpyxl import load_workbook, Workbook

#region read+write xlsx util

"""
wb aka workbook
ws aka worksheet
"""

class ExcelRead:

    def xlrd_read_excel(self, path_file):
        wb = xlrd.open_workbook(path_file)
        ws = wb.sheet_by_index(0)

        raw_headers = [cell.value for cell in ws.row(0)]
        col_count = len(raw_headers)

        data_return = ()
        for rx in range(ws.nrows):
            data_return += ([ws.cell_value(rowx=rx, colx=ci) for ci in range(col_count)], )
        return data_return

    def openpyxl_read_excel(self, path_file):
        wb = load_workbook(path_file)
        ws = wb.active

        max_row = ws.max_row
        max_column = ws.max_column

        data_return = ()
        for i in range(1, max_row + 1):  # iterate over all cells
            row = []
            for j in range(1, max_column + 1):
                cell_obj = ws.cell(row=i, column=j)
                row.append(cell_obj.value)
            data_return += (row,)
        return data_return


class ExcelWrite:

    def xlsxwriter_write_to_excel(self, file_name, worksheet_name, data_to_write=()):
        # Create an new Excel file and add a worksheet.
        wb = xlsxwriter.Workbook(file_name)
        ws = wb.add_worksheet(worksheet_name)  # Defaults to Sheet1

        # Add a bold format to use to highlight cells.
        bold = wb.add_format({'bold': True})

        # Add a number format for cells with money.
        money = wb.add_format({'num_format': '$#,##0'})

        # Some data we want to write to the worksheet.
        expenses = data_to_write

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0
        col = 0
        # Iterate over the data and write it out row by row.
        for item, cost in (expenses):
            ws.write(row, col, item)
            ws.write(row, col + 1, cost)
            row += 1
        wb.close()

    def openpyxl_write_new_excel_file(self, file_name, data_to_write=()):
        wb = Workbook()
        ws = wb.active

        # append all rows
        for row in data_to_write:
            ws.append(tuple(row))

        # save file
        wb.save(file_name)
    
    def openpyxl_write_existing_excel_file(self, file_name, data_to_write=()):
        wb = load_workbook(file_name)
        ws = wb.active
        max_row = ws.max_row
        for row in data_to_write:
            for i in range(1, len(row) + 1):  ## first start column and row are 1
                cell = ws.cell(row=max_row + 1, column=i)
                cell.value = row[i - 1]
            max_row += 1 ## add 1 to write new row
        wb.save(file_name)
        

#endregion read+write xlsx util

class ExcelReadT:

    def xlrd_read_excel(self, path_file):
        wb = xlrd.open_workbook(path_file)
        ws = wb.sheet_by_index(0)

        raw_headers = [cell.value for cell in ws.row(0)]
        col_count = len(raw_headers)

        data_return = ()
        for rx in range(ws.nrows):
            data_return += ([ws.cell_value(rowx=rx, colx=ci) for ci in range(col_count)], )
        return data_return

    def openpyxl_read_excel(self, path_file):
        wb = load_workbook(path_file)
        ws = wb.active

        max_row = ws.max_row
        max_column = ws.max_column

        data_return = ()
        for i in range(1, max_column + 1):  # iterate over all cells
            col = []
            for j in range(1, max_row + 1):
                cell_obj = ws.cell(row=j, column=i)
                col.append(cell_obj.value)
            data_return += (col,)
        return data_return


class ExcelWriteT:

    def xlsxwriter_write_to_excel(self, file_name, worksheet_name, data_to_write=()):
        # Create an new Excel file and add a worksheet.
        wb = xlsxwriter.Workbook(file_name)
        ws = wb.add_worksheet(worksheet_name)  # Defaults to Sheet1

        # Add a bold format to use to highlight cells.
        bold = wb.add_format({'bold': True})

        # Add a number format for cells with money.
        money = wb.add_format({'num_format': '$#,##0'})

        # Some data we want to write to the worksheet.
        expenses = data_to_write

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0
        col = 0
        # Iterate over the data and write it out row by row.
        for item, cost in (expenses):
            ws.write(row, col, item)
            ws.write(row, col + 1, cost)
            row += 1
        wb.close()

    def openpyxl_write_new_excel_file(self, file_name, data_to_write=()):
        wb = Workbook()
        ws = wb.active

        # append all rows
        for row in data_to_write:
            ws.append(tuple(row))

        # save file
        wb.save(file_name)
    
    def openpyxl_write_existing_excel_file(self, file_name, data_to_write=()):
        wb = load_workbook(file_name)
        ws = wb.active
        max_column = ws.max_column
        for col in data_to_write:
            for i in range(1, len(col) + 1):  ## first start column and row are 1
                cell = ws.cell(row=i, column=max_column + 1)
                cell.value = col[i - 1]
            max_column += 1 ## add 1 to write new row
        wb.save(file_name)